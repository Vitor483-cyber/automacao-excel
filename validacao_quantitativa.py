# -- coding: utf-8 --
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def extrair_unidades(desc, palavras_chave):
    if not isinstance(desc, str):
        return None
    desc = desc.lower()
    desc = re.sub(r'\b\d+(\.\d+)?\s*(g|gr|grm|kg|ml|l|lt|litros?|l\u00edquido|liq)\b', '', desc)
    desc = re.sub(r'\bliq\b', '', desc)
    match_par = re.search(r'\(\s*(\d+)\s*[xX\u00d7*]', desc)
    if match_par:
        return int(match_par.group(1))
    for palavra in palavras_chave:
        padrao = rf'\b(\d+)\s*{re.escape(palavra.lower())}\b'
        match = re.search(padrao, desc)
        if match:
            return int(match.group(1))
    return None

def extrair_unid_parenteses(desc):
    if not isinstance(desc, str):
        return None
    desc = re.sub(r'\b\d+(\.\d+)?\s*(g|gr|grm|kg|ml|l|lt|litros?)\b', '', desc.lower())
    match = re.search(r'\(\s*(\d+)\s*[xX\u00d7*]', desc)
    return int(match.group(1)) if match else None

def extrair_unidade_interna(desc):
    if not isinstance(desc, str):
        return None
    desc = re.sub(r'\b\d+(\.\d+)?\s*(g|gr|grm|kg|ml|l|lt|litros?)\b', '', desc.lower())
    match = re.search(r'\(\s*(\d+)\s*[xX\u00d7*]', desc)
    return int(match.group(1)) if match else None

def extrair_unidade_caixa(desc):
    if not isinstance(desc, str):
        return None
    desc = desc.lower()
    match = re.search(r'(?:cx|c)\s*c?[\/]?\s*(\d+)\s*(?:un|und|unid)?', desc)
    return int(match.group(1)) if match else None

def analise_quantitativa(row):
    desc = str(row["#BR LOC 1000001 : ITM_DESC"]).lower()
    total_unid = row.get("UNID_TOTAIS")
    contenido = row.get("#BR LOC 008 : CONTENIDO G2G")
    if pd.notna(total_unid) and total_unid > 0:
        return f"Pack de {int(total_unid)} unidades"
    if any(x in desc for x in ["pack", "x", "un", "und", "unid"]):
        if pd.notna(contenido) and contenido <= 10:
            return f"Possível pack sem número claro (base conteúdo: {int(contenido)})"
        else:
            return "Possível pack sem número claro (0 unidades)"
    if pd.notna(contenido) and int(contenido) == 1:
        return "Unitário (1 unidade)"

def compara_contenido_global(row):
    try:
        return "OK" if int(row.get("#BR LOC 008 : CONTENIDO G2G")) == int(row.get("GLOBAL TOTAL PACKS IN MULTIPACK")) else "DIVERGENTE"
    except:
        return "NÃO IDENTIFICADO"

def comparar_analise_vs_global_pack(row):
    try:
        a = int(re.search(r'\d+', str(row["ANALISE_QUANTITATIVA"])).group())
        g = int(row["GLOBAL TOTAL PACKS IN MULTIPACK ANALISE"])
        return "OK" if a == g else "DIVERGENTE"
    except:
        return "NÃO IDENTIFICADO"

def gerar_resumo(df):
    resumo = df["ANALISE_QUANTITATIVA"].value_counts().reset_index()
    resumo.columns = ["TIPO DE CLASSIFICAÇÃO", "QUANTIDADE"]
    resumo["%"] = round(100 * resumo["QUANTIDADE"] / resumo["QUANTIDADE"].sum(), 1)
    return resumo

def gerar_resumo_total_embalagem(df):
    total_identificados = df["TOTAL_EMBALAGEM"].notna().sum()
    total = len(df)
    return pd.DataFrame({
        "TIPO": ["Com Total Identificado", "Sem Total"],
        "QUANTIDADE": [total_identificados, total - total_identificados],
        "%": [round(100 * total_identificados / total, 1),
              round(100 * (total - total_identificados) / total, 1)]
    })

def aplicar_formatacao_excel(path, colunas):
    wb = load_workbook(path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    for coluna in colunas:
        if coluna not in header: continue
        col_idx = header.index(coluna) + 1
        letra = get_column_letter(col_idx)
        for row in range(2, ws.max_row + 1):
            valor = ws[f"{letra}{row}"].value
            if not valor: continue
            cor = None
            if coluna == "ANALISE_QUANTITATIVA":
                cor = "C6EFCE" if "pack" in valor.lower() else "D9E1F2" if "unitário" in valor.lower() else "FFF2CC" if "possível" in valor.lower() else "F8CBAD"
            elif coluna in ["COMPARA_CONTENIDO_VS_GLOBAL", "COMPARA_ANALISE_VS_GLOBAL_PACK"]:
                cor = "C6EFCE" if "ok" in valor.lower() else "FCE4D6" if "divergente" in valor.lower() else "D9D9D9"
            elif coluna == "TOTAL_EMBALAGEM":
                cor = "E2EFDA"
            if cor:
                ws[f"{letra}{row}"].fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                ws[f"{letra}{row}"].font = Font(bold=True)
    wb.save(path)

def adicionar_resumo_excel(path, df1, df2):
    wb = load_workbook(path)
    for name, df in [("Resumo", df1), ("Totais Identificados", df2)]:
        ws = wb.create_sheet(title=name)
        for r in dataframe_to_rows(df, index=False, header=True): ws.append(r)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    wb.save(path)

def main():
    arquivo_descripciones = r"C:\\Users\\govi5001\\OneDrive - NIQ\\Pictures\\Validation_Tool\\IDE.csv"
    arquivo_palavras = r"C:\\Users\\govi5001\\OneDrive - NIQ\\Documents\\Validation_Tool\\palavras_chave.xlsx"

    try:
        df = pd.read_csv(arquivo_descripciones, low_memory=False)
        df_keywords = pd.read_excel(arquivo_palavras)
        palavras_chave = df_keywords["PALAVRAS CHAVE"].dropna().astype(str).tolist()

        df["GLOBAL TOTAL PACKS IN MULTIPACK ORIG"] = df["GLOBAL TOTAL PACKS IN MULTIPACK"]
        df["GLOBAL TOTAL PACKS IN MULTIPACK ANALISE"] = (
            df["GLOBAL TOTAL PACKS IN MULTIPACK ORIG"].astype(str).str.strip().str.replace(r"[^\d.]", "", regex=True)
        )
        df["GLOBAL TOTAL PACKS IN MULTIPACK ANALISE"] = pd.to_numeric(df["GLOBAL TOTAL PACKS IN MULTIPACK ANALISE"], errors='coerce')
        df["#BR LOC 008 : CONTENIDO G2G"] = pd.to_numeric(df["#BR LOC 008 : CONTENIDO G2G"], errors='coerce')

        df["UNID_ENCONTRADAS"] = df["#BR LOC 1000001 : ITM_DESC"].apply(lambda x: extrair_unidades(x, palavras_chave))
        df["UNID_PARENTESES"] = df["#BR LOC 1000001 : ITM_DESC"].apply(extrair_unid_parenteses)
        df["UNID_TOTAIS"] = df["UNID_ENCONTRADAS"].combine_first(df["UNID_PARENTESES"])

        df["ANALISE_QUANTITATIVA"] = df.apply(analise_quantitativa, axis=1)
        df["COMPARA_CONTENIDO_VS_GLOBAL"] = df.apply(compara_contenido_global, axis=1)
        df["COMPARA_ANALISE_VS_GLOBAL_PACK"] = df.apply(comparar_analise_vs_global_pack, axis=1)

        df["UNIDADE_INTERNA"] = df["#BR LOC 1000001 : ITM_DESC"].apply(extrair_unidade_interna)
        df["UNIDADE_CAIXA"] = df["#BR LOC 1000001 : ITM_DESC"].apply(extrair_unidade_caixa)

        df["TOTAL_EMBALAGEM"] = df.apply(
            lambda row: row["UNIDADE_INTERNA"] * row["UNIDADE_CAIXA"]
            if pd.notna(row["UNIDADE_INTERNA"]) and pd.notna(row["UNIDADE_CAIXA"])
            else None,
            axis=1
        )

        df_resumo = gerar_resumo(df)
        df_total_resumo = gerar_resumo_total_embalagem(df)

        df["GLOBAL TOTAL PACKS IN MULTIPACK"] = df["GLOBAL TOTAL PACKS IN MULTIPACK ORIG"]
        df.drop(columns=["GLOBAL TOTAL PACKS IN MULTIPACK ORIG", "GLOBAL TOTAL PACKS IN MULTIPACK ANALISE"], inplace=True)

        output = os.path.splitext(arquivo_descripciones)[0] + "_analisado.xlsx"
        df.to_excel(output, index=False)

        aplicar_formatacao_excel(output, [
            "ANALISE_QUANTITATIVA", 
            "COMPARA_CONTENIDO_VS_GLOBAL", 
            "COMPARA_ANALISE_VS_GLOBAL_PACK",
            "TOTAL_EMBALAGEM"
        ])

        adicionar_resumo_excel(output, df_resumo, df_total_resumo)

        print(f"\u2705 Análise finalizada com sucesso. Arquivo salvo em:\n{output}")

    except Exception as e:
        print(f"\u274c Erro durante a execução: {e}")

if __name__ == "__main__":
    main()
