# -- coding: utf-8 --
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ---------------------- FUNÇÕES ----------------------

def extrair_unidades(desc, palavras_chave):
    if not isinstance(desc, str):
        return None
    desc = desc.lower()

    match_par = re.search(r'\(\s*(\d+)\s*[xX×*]', desc)
    if match_par:
        return int(match_par.group(1))

    match_sem_paren = re.search(r'(?<!\()\b(\d+)\s*[xX×*]', desc)
    if match_sem_paren:
        return int(match_sem_paren.group(1))

    for palavra in palavras_chave:
        if not palavra:
            continue
        padrao = rf'(\d+)\s*{re.escape(palavra.lower())}\b'
        match_kw = re.search(padrao, desc)
        if match_kw:
            return int(match_kw.group(1))

    return None

def extrair_unid_parenteses(desc):
    if not isinstance(desc, str):
        return None
    desc = desc.lower()
    match = re.search(r'\(\s*(\d+)\s*[xX×*]', desc)
    if match:
        return int(match.group(1))
    return None

def analise_quantitativa(row):
    desc = str(row["#BR LOC 1000001 : ITM_DESC"]).lower() if pd.notna(row["#BR LOC 1000001 : ITM_DESC"]) else ""
    total_unid = row.get("UNID_TOTAIS", None)
    contenido = row.get("#BR LOC 008 : CONTENIDO G2G", None)

    if pd.notna(total_unid) and total_unid > 0:
        return f"Pack de {int(total_unid)} unidades"
    if any(x in desc for x in ["pack", "x", "un", "und", "unid"]):
        if pd.notna(contenido) and contenido > 0:
            return f"Possível pack sem número claro (base conteúdo: {int(contenido)})"
        else:
            return "Possível pack sem número claro (0 unidades)"
    if pd.notna(contenido) and int(contenido) == 1:
        return "Unitário (1 unidade)"
    if pd.notna(total_unid) and int(total_unid) == 1:
        return "Unitário (1 unidade)"
    return "Não identificado (0 unidades)"

def compara_contenido_global(row):
    cont = row.get("#BR LOC 008 : CONTENIDO G2G", None)
    global_unid = row.get("GLOBAL TOTAL ITEMS IN PACK", None)
    try:
        cont_int = int(cont)
        global_int = int(global_unid)
        return "OK" if cont_int == global_int else "DIVERGENTE"
    except (ValueError, TypeError):
        return "NÃO IDENTIFICADO"

def comparar_analise_vs_global_pack(row):
    analise = row.get("ANALISE_QUANTITATIVA", "")
    global_pack = row.get("GLOBAL TOTAL PACKS IN MULTIPACK", "")

    match_analise = re.search(r'(\d+)', str(analise))
    valor_analise = int(match_analise.group(1)) if match_analise else None

    match_global = re.search(r'(\d+)', str(global_pack))
    valor_global = int(match_global.group(1)) if match_global else None

    if valor_analise is not None and valor_global is not None:
        return "OK" if valor_analise == valor_global else "DIVERGENTE"
    return "NÃO IDENTIFICADO"

def gerar_resumo(df):
    resumo = df["ANALISE_QUANTITATIVA"].value_counts().reset_index()
    resumo.columns = ["TIPO DE CLASSIFICAÇÃO", "QUANTIDADE"]
    total = resumo["QUANTIDADE"].sum()
    resumo["%"] = round(100 * resumo["QUANTIDADE"] / total, 1)
    return resumo

def aplicar_formatacao_excel(caminho_excel, colunas_alvo):
    wb = load_workbook(caminho_excel)
    ws = wb.active
    ws.freeze_panes = "A2"
    header = [cell.value for cell in ws[1]]

    for coluna in colunas_alvo:
        if coluna not in header:
            continue
        col_idx = header.index(coluna) + 1
        col_letra = get_column_letter(col_idx)

        for row in range(2, ws.max_row + 1):
            valor = ws[f"{col_letra}{row}"].value
            if not valor:
                continue

            font = Font(bold=True)

            if coluna == "ANALISE_QUANTITATIVA":
                if "pack de" in valor.lower():
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "possível" in valor.lower():
                    fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif "unitário" in valor.lower():
                    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                else:
                    fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

            elif coluna in ["COMPARA_CONTENIDO_VS_GLOBAL", "COMPARA_ANALISE_VS_GLOBAL_PACK"]:
                if "ok" in valor.lower():
                    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "divergente" in valor.lower():
                    fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                else:
                    fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            ws[f"{col_letra}{row}"].fill = fill
            ws[f"{col_letra}{row}"].font = font

    wb.save(caminho_excel)

def adicionar_resumo_excel(caminho_excel, df_resumo):
    wb = load_workbook(caminho_excel)
    ws_resumo = wb.create_sheet(title="Resumo")

    for r in dataframe_to_rows(df_resumo, index=False, header=True):
        ws_resumo.append(r)

    for cell in ws_resumo[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    wb.save(caminho_excel)

# ---------------------- MAIN ----------------------

def main():
    arquivo_descripciones = r"C:\Users\govi5001\OneDrive - NIQ\Pictures\Validation_Tool\IDE.csv"
    arquivo_palavras = r"C:\Users\govi5001\OneDrive - NIQ\Documents\Validation_Tool\palavras_chave.xlsx"

    try:
        df = pd.read_csv(arquivo_descripciones, low_memory=False)
        df_keywords = pd.read_excel(arquivo_palavras)
        palavras_chave = df_keywords["PALAVRAS CHAVE"].dropna().astype(str).tolist()

        df["UNID_ENCONTRADAS"] = df["#BR LOC 1000001 : ITM_DESC"].apply(lambda x: extrair_unidades(x, palavras_chave))
        df["UNID_PARENTESES"] = df["#BR LOC 1000001 : ITM_DESC"].apply(extrair_unid_parenteses)
        df["UNID_TOTAIS"] = df["UNID_ENCONTRADAS"].combine_first(df["UNID_PARENTESES"])

        df["#BR LOC 008 : CONTENIDO G2G"] = pd.to_numeric(df["#BR LOC 008 : CONTENIDO G2G"], errors='coerce')
        df["GLOBAL TOTAL ITEMS IN PACK"] = pd.to_numeric(df["GLOBAL TOTAL ITEMS IN PACK"], errors='coerce')

        df["ANALISE_QUANTITATIVA"] = df.apply(analise_quantitativa, axis=1)
        df["COMPARA_CONTENIDO_VS_GLOBAL"] = df.apply(compara_contenido_global, axis=1)
        df["COMPARA_ANALISE_VS_GLOBAL_PACK"] = df.apply(comparar_analise_vs_global_pack, axis=1)

        df_resumo = gerar_resumo(df)

        output_path = os.path.splitext(arquivo_descripciones)[0] + "_analisado.xlsx"
        df.to_excel(output_path, index=False)

        aplicar_formatacao_excel(output_path, [
            "ANALISE_QUANTITATIVA", 
            "COMPARA_CONTENIDO_VS_GLOBAL", 
            "COMPARA_ANALISE_VS_GLOBAL_PACK"
        ])
        adicionar_resumo_excel(output_path, df_resumo)

        print(f"✅ Análise finalizada com sucesso. Arquivo salvo em:\n{output_path}")

    except Exception as e:
        print(f"❌ Erro durante a execução: {e}")

# ---------------------- EXECUTAR ----------------------
if __name__ == "__main__":
    main()