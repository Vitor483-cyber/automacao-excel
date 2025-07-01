import sys
from openpyxl import load_workbook

arquivo_origem = r'C:\Users\govi5001\OneDrive - NIQ\MARKET TRACK BR - Descritivo pobre\CP 51\5102 - CAMILA.xlsx'
arquivo_destino = r'C:\Users\govi5001\OneDrive - NIQ\Documents\teste_openpyxl.xlsx'

# ⚠️ data_only=True → Lê o resultado das fórmulas (não a fórmula em si)
wb_origem = load_workbook(arquivo_origem, data_only=True)
wb_destino = load_workbook(arquivo_destino)

aba_origem = wb_origem['7.Planilha de Análise']
aba_destino = wb_destino.active

# Remover filtro
if aba_origem.auto_filter:
    aba_origem.auto_filter.ref = None
    print("🧹 Filtros da planilha de origem removidos.")

# Detectar linha de cabeçalho
linha_cabecalho = None
for linha in [2, 4]:
    if aba_origem.cell(row=linha, column=1).value:
        linha_cabecalho = linha
        break
if linha_cabecalho is None:
    print("❌ Não foi possível encontrar a linha de cabeçalhos.")
    sys.exit()

print(f"📌 Cabeçalhos encontrados na linha {linha_cabecalho}.")

# Mapear cabeçalhos da origem
cabecalhos_origem = {}
for col in range(1, aba_origem.max_column + 1):
    nome = aba_origem.cell(row=linha_cabecalho, column=col).value
    if nome:
        cabecalhos_origem[nome.strip().lower()] = col  # normaliza para lowercase

print("Cabeçalhos encontrados na origem:", list(cabecalhos_origem.keys()))

# Mapear cabeçalhos da planilha de destino
cabecalhos_destino = {}
for col in range(1, aba_destino.max_column + 1):
    nome = aba_destino.cell(row=1, column=col).value
    if nome:
        cabecalhos_destino[nome.strip().lower()] = col

# Entrada dos cabeçalhos desejados
entrada = input("Digite os cabeçalhos a copiar, separados por vírgula (ex: Item Code, European NAN KEY): ")
lista_cabecalhos = [x.strip().lower() for x in entrada.split(",")]

# Adiciona 'status' automaticamente se existir
if 'status' not in cabecalhos_origem:
    print("❌ Cabeçalho 'status' não encontrado na planilha de origem. Não será possível filtrar.")
    sys.exit()
elif 'status' not in lista_cabecalhos:
    lista_cabecalhos.append('status')

col_status = cabecalhos_origem['status']

# Filtra linhas com status "CORRIGIR"
linhas_validas = []
for linha in range(linha_cabecalho + 1, aba_origem.max_row + 1):
    valor_status = aba_origem.cell(row=linha, column=col_status).value
    if valor_status and str(valor_status).strip().upper() == 'CORRIGIR':
        linhas_validas.append(linha)

print(f"🔎 {len(linhas_validas)} linhas encontradas com Status='CORRIGIR'.")

linha_destino = 2

# Copiar os dados
for cab in lista_cabecalhos:
    if cab not in cabecalhos_origem:
        print(f"⚠️ Cabeçalho '{cab}' não encontrado na origem.")
        continue
    if cab not in cabecalhos_destino:
        print(f"⚠️ Cabeçalho '{cab}' não encontrado na destino.")
        continue

    col_origem = cabecalhos_origem[cab]
    col_destino = cabecalhos_destino[cab]

    for i, linha_origem in enumerate(linhas_validas):
        valor = aba_origem.cell(row=linha_origem, column=col_origem).value
        aba_destino.cell(row=linha_destino + i, column=col_destino).value = valor

    print(f"✅ Cabeçalho '{cab}' copiado com sucesso.")

# Salvar a planilha destino
wb_destino.save(arquivo_destino)
print("✔️ Processamento finalizado com sucesso.")

