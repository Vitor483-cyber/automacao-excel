import sys
from openpyxl import load_workbook

arquivo_origem = r'C:\Users\govi5001\OneDrive - NIQ\MARKET TRACK BR - Descritivo pobre\CP 08\802 INSETICIDAS DOMESTICOS ELETRICOS E REFIL.xlsm'
arquivo_destino = r'C:\Users\govi5001\OneDrive - NIQ\Documents\teste_openpyxl.xlsx'

wb_origem = load_workbook(arquivo_origem)
wb_destino = load_workbook(arquivo_destino)

aba_origem = wb_origem['FORMULAS']
aba_destino = wb_destino.active

# Remove filtros da origem, se existirem
if aba_origem.auto_filter:
    aba_origem.auto_filter.ref = None
    print("🧹 Filtros da planilha de origem removidos.")

# Detecta linha do cabeçalho (exemplo entre linha 2 e 4)
linha_cabecalho = None
for linha in [2, 4]:
    if aba_origem.cell(row=linha, column=1).value:
        linha_cabecalho = linha
        break

if linha_cabecalho is None:
    print("❌ Não foi possível encontrar a linha de cabeçalhos.")
    sys.exit()

print(f"📌 Cabeçalhos encontrados na linha {linha_cabecalho}.")

# Mapeia cabeçalhos da origem (normalizado)
cabecalhos_origem = {}
for col in range(1, aba_origem.max_column + 1):
    nome = aba_origem.cell(row=linha_cabecalho, column=col).value
    if nome:
        cabecalhos_origem[nome.strip().lower()] = col

print("Cabeçalhos encontrados na origem:", list(cabecalhos_origem.keys()))

# Mapeia cabeçalhos da destino (normalizado)
cabecalhos_destino = {}
print("Cabeçalhos encontrados na planilha destino:")
for col in range(1, aba_destino.max_column + 1):
    nome = aba_destino.cell(row=1, column=col).value
    if nome:
        nome_norm = nome.strip().lower()
        cabecalhos_destino[nome_norm] = col
        print(f"Col {col}: '{nome}'")

# Entrada do usuário para cabeçalhos a copiar
entrada = input("Digite os cabeçalhos a copiar, separados por vírgula (ex: item code, european nan key): ")
lista_cabecalhos = [x.strip().lower() for x in entrada.split(",")]

# Verifica se existe o 'status' para filtrar
if 'status' not in cabecalhos_origem:
    print("❌ Cabeçalho 'status' não encontrado na planilha de origem. Não será possível filtrar.")
    sys.exit()

# Garante que 'status' esteja na lista para copiar
if 'status' not in lista_cabecalhos:
    lista_cabecalhos.append('status')

col_status = cabecalhos_origem['status']

# Encontra linhas onde status == 'corrigir' (case insensitive)
linhas_validas = []
for linha in range(linha_cabecalho + 1, aba_origem.max_row + 1):
    valor_status = aba_origem.cell(row=linha, column=col_status).value
    if valor_status and str(valor_status).strip().lower() == 'corrigir':
        linhas_validas.append(linha)

print(f"🔎 {len(linhas_validas)} linhas encontradas com Status='Corrigir'.")

linha_destino = 2  # Começa na linha 2 da planilha destino para colar dados

# Copia os dados das colunas selecionadas
for cab in lista_cabecalhos:
    if cab not in cabecalhos_origem:
        print(f"⚠️ Cabeçalho '{cab}' não encontrado na origem. Pulando...")
        continue
    if cab not in cabecalhos_destino:
        print(f"⚠️ Cabeçalho '{cab}' não encontrado na destino. Pulando...")
        continue

    col_origem = cabecalhos_origem[cab]
    col_destino = cabecalhos_destino[cab]

    for i, linha_origem in enumerate(linhas_validas):
        valor = aba_origem.cell(row=linha_origem, column=col_origem).value
        aba_destino.cell(row=linha_destino + i, column=col_destino).value = valor

    print(f"✅ Cabeçalho '{cab}' copiado com sucesso.")

wb_destino.save(arquivo_destino)
print("✔️ Processamento finalizado com sucesso.")

